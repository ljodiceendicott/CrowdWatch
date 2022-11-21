##REFERENCE:
##https://realpython.com/openpyxl-excel-spreadsheets-python/
from openpyxl import Workbook, load_workbook
import json

# Reading from an excel file
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# read_workbook = load_workbook(filename = "hello_wrld.xlsx")
# print(read_workbook.sheetnames)
# ways to retrieve data like this :
# sheet["A1"] >> <cell 'sheet 1.A1> (Main Cell obj)
# sheet["A1"].value >> the value found in the cell
#ITERATING THROUGH THE DATA
#sheet["A1:C2"]:
# ((<Cell 'Sheet 1'.A1>, <Cell 'Sheet 1'.B1>, <Cell 'Sheet 1'.C1>),
#  (<Cell 'Sheet 1'.A2>, <Cell 'Sheet 1'.B2>, <Cell 'Sheet 1'.C2>))
#sheet["A"] >> entire column can do the same thing with rows



# Writing to a excel file
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# workbook = Workbook()
# sheet = workbook.active
# sheet["A1"] = "Hello"
# sheet["B1"] = "World"

# workbook.save(filename="hello_wrld.xlsx")
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~