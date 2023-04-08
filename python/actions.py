# Luke Jodice
# Actions.py
# This file is where common functions are located. These actions are used throughout the other files
# they are in one location as there are different uses for them and they are very generic allowing for
# them to be used for different pourposes

# checklist
# [X] spreadsheet
# [X] set to json
# [X] Read from Arduino
# [ ] Networking-GET/PUT/POST

import json, sqlite3, time
from datetime import date, datetime, timedelta
from timeloop import Timeloop
from openpyxl import Workbook, load_workbook
import requests

# ////////////////////////////////////////////////////////////////////////////////////////
# Json
# ////////////////////////////////////////////////////////////////////////////////////////

# class needed for object to json transfer
class expJson(object):
    def __init__(self, object):
        self.account = object


# returns data in form of json for a complex object
def serialize(object):
    data = json.dumps(object, default=lambda o: o.__dict__)
    return data


# Returns the data that is found in Json File that is for the specific business returning the business OBJ
# @param - name - Name the json file that is being looked for in user files
def read_from_Json(name):

    # for testing

    # For real imp
    file = open("../UserFiles/" + name + ".json")
    data = json.load(file)
    file.close()
    return data


# This is a Use case from reading in from json
# ------how to access the location name and adding the count member and making it zero------
# data = read_from_Json("Test")
# for i in data['account']['locations']:
#     name = i['name']
#     i['count'] = 0


# Writes the current Data to the json file. that will be updated
# @param data - data to be written to the file -Type(Json Object)
# @param name - name of the file to be written -Type(String)
def write_to_Json(data, name):
    with open(name + ".json", "w") as outfile:
        outfile.write(data)


# Use Case
# adjusting the count of a location then writing to a file
# data = read_from_Json('paddy')
# locations = data['account']['locations']
# loc1= locations[0]
# loc1count = locations[0]['count']+1
# loc1.update({"count": loc1count})
# write_to_Json(serialize(data),'paddy')


# Returns the Date as a string of numbers
# EX.) 2301311802
# order of num => Year-Month-Day-Hour-Minute
def getDate():
    now = datetime.now()
    dt_string = now.strftime("%y-%m-%d %H:%M")
    return dt_string


# ////////////////////////////////////////////////////////////////////////////////////////
# Spreadsheet Code
# ////////////////////////////////////////////////////////////////////////////////////////


# Returns instance of Workbook
# @param name - Name of the business that you are tryin to get access to their spreadsheet
def workbookInstance(name):
    workbook = load_workbook("../UserFiles/" + name + "businessAnalysis.xlsx")
    print("instance made")
    return workbook


# adds data to spreadsheet
# @param fn - File name that will be written to
# @param data- data to be added. How it should be formatted: [get_Date(),count1,count2,count3, ...., countn]
def spreadsheet_Data_Add(fn, data):
    workbook = workbookInstance(fn)
    sheet = workbook.active
    column = nextColumn(fn, workbook)
    print(column)
    for i in range(
        0, len(data)
    ):  # Loops through all of the data that is given from update
        print(data[i])
        # cell = sheet.cell(row=i + 1, column=column)
        print(cell.value)
        # cell.value = data[i]
    workbook.save("../UserFiles/" + fn + "businessAnalysis.xlsx")


# Returns the index of the next avaliable column to be written to
# @param fn - File name that will be written to
# @param wb - Instance of the Workbook
def nextColumn(fn, wb):
    workbook = wb
    sheet = workbook.active
    # loops through all of the columns known
    for col in range(1, sheet.max_column + 2):
        # if the header is not there then the column can be written to
        if (
            sheet.cell(row=1, column=col).value == None
            or sheet.cell(row=1, column=col).value == ""
        ):
            print("found empty spot @", col)
            return col


# ////////////////////////////////////////////////////////////////////////////////////////
# SQL DATABASE
# ////////////////////////////////////////////////////////////////////////////////////////


# Returns SQL DB object as a Tuple
# Tuple[0] - sqliteConnection
# Tuple[1] - Cursor obj
# @param name - name of the sql db
def databaseConnect(name):
    sqliteConnection = sqlite3.connect("../UserFiles/" + name + "SQL.db")
    cursor = sqliteConnection.cursor()
    return sqliteConnection, cursor


def EnterDatabase(obj, data, table):
    db = databaseConnect(table)
    for i in range(len(data - 2)):
        idx = i + 2
        print()


# ////////////////////////////////////////////////////////////////////////////////////////
# Networking
# ////////////////////////////////////////////////////////////////////////////////////////

# ////////////////////////////////////////////////////////////////////////////////////////
# Networking
# ////////////////////////////////////////////////////////////////////////////////////////

# This will be the inital communication from python
# loading setting the values in the db
def getRequest():
    # Replace this with the host website when a website is hosted
    url = "http://localhost:3000"
    print()


def putRequest(locationObj):
    r = requests.put(
        "http://localhost:3000/putReq",
        data={name: locationObj.name, count: locationObj.count},
    )
    # success code
    print(r)
    # Content of Req
    print(r.content)


def putRequest(location):
    r = requests.put("http://localhost:3000/putReq", data={})
    # success code
    print(r)
    # Content of Req


def postRequest():
    # URL for testing
    url = "http://localhost:3000/test"
    # URL for live -- To be added
    data = {"status": "connected", "date": getDate()}
    postlog(data)
    x = requests.post(url, json=data)

    print(x.text)


def postRequest(data):
    # URL for testing
    url = "http://localhost:3000/test"
    # URL for live -- To be added
    data = {"status": "connected", "data": data}
    postlog(data)
    x = requests.post(url, json=data)

    # print(x.text)

# postRequest(locations[0])
# postRequest(history)


def postlog(data):
    url = "http://localhost:3000/logs"
    data = data
    x = requests.post(url, json=data)
    print(x.text)


##~~~Testing Area~~~~##
# print(read_from_Json("paddyFull"))
data = read_from_Json("paddyFull")
locations = data["account"]["locations"]
history = locations[0]["history"]


# locations[0]["count"] = locations[0]["count"] + 1
# print(locations[0])
# print(locations)
# for location in locations:
#     print(location['name'])
# print('reading from file:\n\n\n')


##~~Testing Area End ~~~~~##

# postRequest()


# nextColumn('paddy')
# Sample spreadsheet_Data_Add(NAME_OF_BUS, DATA_FROM_BUS)
# Sample call: spreadsheet_Data_Add(account.name, account.data)
# spreadsheet_Data_Add('paddy',[get_Date(),1,2,3])
# print(serialize(data))


# file = open("..\Website\PeopleWatchSite\public\data\data.json.json")
# data = json.load(file)
# dataloc = data[3]
# dataUpdate = data[3]['amount']
# dataUpdate = dataUpdate + 3
# dataloc.update({"amount":dataUpdate})
# write_to_Json(serialize(data), '..\Website\PeopleWatchSite\public\data\data.json')
# locations = data['account']['locations']
# loc1= locations[0]
# loc1count = locations[0]['count']+1
# loc1.update({"count": loc1count})
# write_to_Json(serialize(data),'./UserFiles/paddy')
