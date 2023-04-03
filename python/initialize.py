import sys
from openpyxl import Workbook, load_workbook
import json
import sqlite3


class Account:
    def __init__(self, name, locations):
        self.name = name
        self.locations = locations

    def __str__(self):
        return self.name

    # def jsonRep(self):
    #     return dict(LocName= self.name, locations=self.locations)
    def getLocationName(self, i):
        return self.locations[i].name

    def numlocations(self):
        return len(self.locations)

    def filecreation(self):
        spreadsheetSetup(self.locations, Account.name)
        print("file Creation")

    def getLocations(self):  # This function should return the locations
        loc = []


class Location:
    def __init__(self, name, maxCap):
        self.name = name
        self.maxCap = maxCap

    # def jsonRep(self):
    #     return dict(name= self.name, maxCap= self.maxCap)
    def __str__(self):
        return self.name


class expJson(object):
    def __init__(self, object):
        self.account = object


def serialize(object):
    data = json.dumps(object, default=lambda o: o.__dict__)
    return data


def accountManualInput():
    # Make user
    businessname = input("What is the name of the business?")
    num_loc = int(input("how many locations are trying to be kept track of?"))
    locations = []
    for i in range(num_loc):
        name = input("What is the nickname for this location?")
        maxcap = int(input("what is the maximum capacity for this location?"))
        loc = Location(name, maxcap)
        locations.append(loc)
    return Account(businessname, locations)


def accountInputData(path):  # This is the input type for if you are using a file
    with open(path, "r") as f:
        business_Name = f.readline()
        business_Name = business_Name[0 : len(business_Name) - 1]
        section_Amount = f.readline()
        section_Amount = int(section_Amount[0])
        locations = []
        for i in range(section_Amount):
            section_Name = f.readline()
            section_Max_Cap = f.readline()
            location = Location(section_Name.strip(), int(section_Max_Cap.strip()))
            locations.append(location)
        return Account(business_Name, locations)


def spreadsheetSetup(account):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Section Name"
    sheet["B1"] = "Max Capacity"
    eof = len(account.locations)
    for i in range(eof):
        print(str(account.locations[i].name))
        print(int(account.locations[i].maxCap))
        if i == 0:
            print()
        else:
            sheet.cell(row=i + 1, column=1).value = account.locations[i].name
            sheet.cell(row=i + 1, column=2).value = account.locations[i].maxCap
    workbook = workbook.save("../UserFiles/" + account.name + "businessAnalysis.xlsx")
    print("Workbook created for " + account.name)


def databaseSetup(account):
    sqliteConnection = sqlite3.connect("../UserFiles/" + account.name + "SQL.db")
    cursor = sqliteConnection.cursor()

    sql_command = """
        CREATE TABLE TEST(
            locID,
            name,
            location,
        );
        """
    # cursor.execute(sql_command)
    # cursor.execute("CREATE TABLE TEST(locID,name,location);")

    # cursor.commit()
    # sqliteConnection.close()
    location = """
        CREATE TABLE Location(
            LocationName VARCHAR(20) PRIMARY KEY,
            MaxCapacity INTEGER,
            AverageCapacity INTEGER
        );
        """
    locationSum = """
        CREATE TABLE LocSummary(
            LocationTimeStamp VARCHAR(32) PRIMARY KEY,
            LocationName VARCHAR(20) FORIEGN KEY,
            TimeStamp VARCHAR(12)
        );
        """
    business = """
        CREATE TABLE Business(
             BusinessName VARCHAR(20) PRIMARY KEY
         );
        """
    business_insert = """
        INSERT INTO Business VALUES ("name")  
         """
    # cursor.execute(location)
    # cursor.execute(locationSum)
    # cursor.execute(business)
    cursor.execute(business_insert)
    sqliteConnection.commit()

    print("SQL command executed")


account = object()
if len(sys.argv) < 2:  # This means that the user does not have a file
    with open("../Documentation/intro.txt", "r") as f:
        contents = f.read()
        print(contents)
        account = accountManualInput()
elif (
    len((sys.argv)) > 1
):  # This means that the user has a file that they would like to pull data from
    # print(sys.argv[1])
    path = sys.argv[1]
    account = accountInputData(path)
    print(account)

spreadsheetSetup(account)  # This is what is used to add it to the spreadsheet)
exp_Json = expJson(account)
print(serialize(exp_Json))  # prints out Json --->>>should be exported into a file
json_to_be_exported = serialize(exp_Json)
# databaseSetup(account) #This writes up the inital SQL file
